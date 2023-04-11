from fastapi import HTTPException
from database.query import query_get, query_put, query_update, query_get_timecard
from .auth import Auth
from .models import *

from datetime import timedelta, datetime

from collections import defaultdict
from typing import List, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import pytz

# Set the download directory
DOWNLOAD_DIR = "./downloads/"

auth_handler = Auth()

###################################################################################################
############################################# USER ################################################
###################################################################################################

# register user using UserSignUpRequestModel
def register_user(user_model: UserSignUpRequestModel):
    user = get_user_by_username(user_model.user_name)
    if len(user) != 0:
        raise HTTPException(
            status_code=409, detail='User Name already exists.')
    hashed_password = auth_handler.encode_password(user_model.user_password)

    query_put("""
                INSERT INTO user (
                    user_first_name,
                    user_last_name,
                    user_name,
                    user_password_hash,
                    user_timezone,
                    user_role
                ) VALUES (%s,%s,%s,%s,%s,%s)
                """,
            (
                user_model.user_first_name,
                user_model.user_last_name,
                user_model.user_name,
                hashed_password,
                user_model.user_timezone,
                user_model.user_role
            )
            )
    user = get_user_by_username(user_model.user_name)
    return user[0]

# sign in user using SignInRequestModel
def sign_in_user(user_model: SignInRequestModel):
    user = get_user_by_username(user_model.user_name)
    if len(user) == 0:
        raise HTTPException(
            status_code=404, detail='User Name user not found.')
    if not auth_handler.verify_password(user_model.user_password, user[0]['user_password_hash']):
        raise HTTPException(
            status_code=401, detail='Incorrect password.')
    user = get_user_by_username(user_model.user_name)
    return user[0]

"""
def signin_user(email, password):
    user_dict = get_user_by_email(email)
    if len(user_dict) == 0:
        print('Invalid email')
        raise HTTPException(status_code=401, detail='Invalid email')
    if (not auth_handler.verify_password(password, user_dict[0]['password_hash'])):
        print('Invalid password')
        raise HTTPException(status_code=401, detail='Invalid password')
    return user_dict[0]
"""


####################################################################################################
######################################## USER DATA QUERIES #########################################
####################################################################################################

# get all users query
def get_all_users():
    user = query_get("""
        SELECT  
            user.id,
            user.user_first_name,
            user.user_last_name,
            user.user_name
        FROM user
        """, ())
    return user

# get user by username query
def get_user_by_username(user_name: str):
    user = query_get("""
        SELECT 
            user.id,
            user.user_first_name,
            user.user_last_name,
            user.user_name,
            user.user_password_hash,
            user.user_timezone
        FROM user 
        WHERE user_name = %s
        """, (user_name))
    return user

# get user by id query
def get_user_by_id(id: int):
    user = query_get("""
        SELECT 
            user.id,
            user.user_first_name,
            user.user_last_name,
            user.user_name,
        FROM user 
        WHERE id = %s
        """, (id))
    return user


###################################################################################################
######################################## USER TIME SHEET ##########################################
###################################################################################################

# get user timesheet data by username query

def get_user_timesheet_by_username(user_name: str) -> List[Dict]:
    user = query_get("""
        SELECT 
            clock_times.clock_in_time,
            clock_times.clock_out_time,
            user.user_timezone
        FROM user
        LEFT JOIN clock_times ON user.id = clock_times.user_id 
        WHERE user_name = %s
        """, (user_name,))
    
    try:
        data = []
        for row in user:
            clock_in_time = row['clock_in_time']
            clock_out_time = row['clock_out_time']
            user_timezone = row['user_timezone']
            data.append({'clock_in_time': clock_in_time, 'clock_out_time': clock_out_time, 'user_timezone': user_timezone})

        # Step 1: Sort the list by the clock_in_time
        sorted_data = sorted(data, key=lambda k: k['clock_in_time'])

        # Step 2: Calculate the total hours for each day
        days_data = {}
        for item in sorted_data:
            clock_in_time = item['clock_in_time']
            clock_out_time = item['clock_out_time']
            user_timezone = pytz.timezone(item['user_timezone'])
            clock_in_time = user_timezone.localize(clock_in_time)
            clock_out_time = user_timezone.localize(clock_out_time)
            date = clock_in_time.date()
            if date not in days_data:
                days_data[date] = {'clock_in_times': [clock_in_time.strftime('%I:%M:%S %p')], 'clock_out_times': [clock_out_time.strftime('%I:%M:%S %p')], 'total_hours': 0}
            else:
                days_data[date]['clock_in_times'].append(clock_in_time.strftime('%I:%M:%S %p'))
                days_data[date]['clock_out_times'].append(clock_out_time.strftime('%I:%M:%S %p'))

            days_data[date]['total_hours'] += (clock_out_time - clock_in_time).total_seconds() / 3600.0

        # Step 3: Create a new key that holds the week number
        for i, (key, value) in enumerate(sorted(days_data.items()), 1):
            week_number = (i-1) // 7 + 1
            days_data[key]['week_number'] = week_number

        # Step 4: Calculate the total hours for each week
        weeks_data = {}
        for key, value in days_data.items():
            week_number = value['week_number']
            if week_number not in weeks_data:
                weeks_data[week_number] = {'total_hours': 0}
            weeks_data[week_number]['total_hours'] += value['total_hours']

        # Create a list of the formatted results
        results = []
        for key, value in days_data.items():
            week_number = value['week_number']
            results.append({
                'date': key.strftime('%Y-%m-%d'),
                'clock_in_times': value['clock_in_times'],
                'clock_out_times': value['clock_out_times'],
                'total_hours': round(value['total_hours'], 2),
                'week_total_hours': round(weeks_data[week_number]['total_hours'], 2),
                'user_timezone': item['user_timezone']
            })

        return results
    except Exception as e:
        print(e)
        return []





# admin function to get all users time records for the week and month and return a list of the formatted results
def get_all_users_time_records():
    # get all users time records
    data = query_get("""
        SELECT 
            clock_times.id,
            clock_times.user_id,
            clock_times.clock_in_time,
            clock_times.clock_out_time,
            user.user_name,
            user.user_timezone
        FROM clock_times
        INNER JOIN user ON clock_times.user_id = user.id
        ORDER BY clock_times.clock_in_time ASC;
    """)

    # convert the user dictionary to a UserResponseModel object
    user = UserTimeZoneResponseModel(**data)

    # get the timezone of the user
    set_timezone = pytz.timezone(user.user_timezone)

    # convert the clock_in_time and clock_out_time to the user's timezone
    clock_in = data['clock_in_time'].astimezone(set_timezone)
    clock_out = data['clock_out_time'].astimezone(set_timezone)

    # convert the clock_in_time and clock_out_time to a string
    clock_in = clock_in.strftime('%Y-%m-%d %H:%M:%S')
    clock_out = clock_out.strftime('%Y-%m-%d %H:%M:%S')

    # create a list of the formatted results
    results = []
    for item in data:
        results.append({
            'id': item['id'],
            'user_id': item['user_id'],
            'user_name': item['user_name'],
            'clock_in_time': clock_in,
            'clock_out_time': clock_out,
            'total_hours': round((item['clock_out_time'] - item['clock_in_time']).total_seconds() / 3600.0, 2),
        })

    return results



###################################################################################################
######################################## USER TIME QUERIES ########################################
###################################################################################################

# get users timezone by username and return only the timezone data
def get_user_timezone_by_username(user_name: str):
    user = query_get("""
        SELECT 
            user.id,
            user.user_name,
            user.user_timezone
        FROM user 
        WHERE user_name = %s
        """, (user_name))
    return user[0]

# save the user's clock-in time to the database
def save_user_time_by_username_clockin(user_name: str):
    # get the user object from the user's username
    user_dict = get_user_timezone_by_username(user_name)
    user_id = user_dict['id']

    # convert the user dictionary to a UserResponseModel object
    user = UserTimeZoneResponseModel(**user_dict)

    # get the timezone of the user
    set_timezone = pytz.timezone(user.user_timezone)

    # check if the user has an existing clock-in record without a clock-out time
    existing_record = query_get("""
        SELECT * FROM clock_times WHERE user_id = %s AND clock_out_time IS NULL;
    """, (user_id))

    #check if the user has any time records
    if len(existing_record) == 0:
            
        # get the current time in the user's timezone
        clock_in = datetime.now(set_timezone)

        # save the new clock-in time to the database
        query_put("""
            INSERT INTO clock_times (user_id, clock_in_time, clock_out_time)
            VALUES (%s, %s, NULL);
        """,
            (user_id, clock_in)
        )

        return clock_in

    elif existing_record is not None:
        # user has an existing clock-in record without a clock-out time, return an error message
        return "Error: You cannot clock in again until you have clocked out."

# save the user's clock-out time to the database
def save_user_time_by_username_clockout(user_name: str):
    # get the user object from the user's username
    user_dict = get_user_timezone_by_username(user_name)
    user_id = user_dict['id']

    # convert the user dictionary to a UserResponseModel object
    user = UserTimeZoneResponseModel(**user_dict)

    # get the timezone of the user
    set_timezone = pytz.timezone(user.user_timezone)

    # get the current time in the user's timezone
    clock_out = datetime.now(set_timezone)

    # if there is no clock in time for the user, return an error message
    if len(query_get("""
        SELECT * FROM clock_times WHERE user_id = %s AND clock_out_time IS NULL;
    """, (user_id))) == 0:
        return "Error: You cannot clock out without first clocking in."

    # update the clock-out time for the most recent clock-in record for the user
    query_put("""
        UPDATE clock_times
        SET clock_out_time = %s
        WHERE user_id = %s AND clock_out_time IS NULL
        ORDER BY clock_in_time DESC
        LIMIT 1;
    """,
        (clock_out, user_id)
    )

    return clock_out


###################################################################################################
############################################## ADMIN ##############################################
###################################################################################################

# admin function to return admin role
def admin_get_role(admin_user: str):
    admin = get_admin_by_name(admin_user)
    return admin[0]['admin_role']


# get admin data by username using AdminResponseModel and return only the admin data
def get_admin_by_name(admin_user: str):
    admin = query_get("""
        SELECT 
            admin.id,
            admin.admin_user,
            admin.admin_password_hash,
            admin.admin_timezone,
            admin.admin_role
        FROM admin 
        WHERE admin_user = %s
        """, (admin_user))
    return admin

# admin function to signup admin user using AdminSignUpRequestModel, if admin has already been created, return an error message
def admin_signup(admin_model: AdminSignUpRequestModel):
    # check if admin has already been created
    admin = get_admin_by_name(admin_model.admin_user)

    if len(admin) != 0:
        raise HTTPException(
            status_code=409, detail='Admin user already exists.')

    # hash the password using AuthHandler
    hashed_password = auth_handler.encode_password(admin_model.admin_password)

    # save the new admin name, first name,  last name,  password , role to the database
    query_put("""
        INSERT INTO admin (
            admin_user,
            admin_password_hash,
            admin_first_name,
            admin_last_name,
            admin_timezone,
            admin_role
            ) VALUES (%s, %s, %s, %s, %s, %s);
    """,
        (
        admin_model.admin_user,
        hashed_password,
        admin_model.admin_first_name,
        admin_model.admin_last_name,
        admin_model.admin_timezone,
        admin_model.admin_role
        )
    )
    admin = get_admin_by_name(admin_model.admin_user)

    return admin[0]



# admin function to login admin user using AdminSignInRequestModel, if admin has not been created, return an error message
def admin_login(admin_model: AdminSignInRequestModel):

    admin = get_admin_by_name(admin_model.admin_user)
    if len(admin) == 0:
        raise HTTPException(
            status_code=404, detail='Admin user not found.')
    if not auth_handler.verify_password(admin_model.admin_password, admin[0]['admin_password_hash']):
        raise HTTPException(
            status_code=401, detail='Incorrect password.')
    admin = get_admin_by_name(admin_model.admin_user)
    return admin[0]


# admin function to create / add non admin users using AdminUserSignUpRequestModel only if the admin has admin role
def admin_create_user(admin: str, user: AdminUserSignUpRequestModel):
    admin = admin_get_role(admin.admin_user)
    # if admin is not an admin, return an error message
    if admin != 'admin':
        return "Error: You do not have permission to view this page."

    # hash the password using AuthHandler
    hashed_password = auth_handler.get_password_hash(user.user_password)

    # save the new user name, first name,  last name,  password , role , timezone to the database
    query_put("""
        INSERT INTO user (user_name, user_first_name, user_last_name, user_password_hash, user_role, user_timezone)
        VALUES (%s, %s, %s, %s, %s, %s);
    """,
        (user.user_name, user.user_first_name, user.user_last_name, hashed_password, user.user_role, user.user_timezone)
    )
    user = get_user_by_username(user.user_name)
    return user[0]

# admin function to delete users using UserUpdateRequestModel only if the admin has admin role
def admin_delete_user(admin: str, user: AdminUserUpdateRequestModel):
    admin = admin_get_role(admin.admin_user)
    # if admin is not an admin, return an error message
    if admin[0]['admin_role'] != 'admin':
        return "Error: You do not have permission to view this page."

    # delete the user from the database
    query_put("""
        DELETE FROM user
        WHERE user_name = %s;
    """,
        (user.user_name)
    )
    return "User deleted."

# admin function to update users password using UserUpdateRequestModel only if the admin has admin role
def admin_update_user_password(admin: str, user: AdminUserUpdateRequestModel):
    admin = admin_get_role(admin.admin_user)
    # if admin is not an admin, return an error message
    if admin != 'admin':
        return "Error: You do not have permission to view this page."

    # hash the password using AuthHandler
    hashed_password = auth_handler.get_password_hash(user.user_password)

    # update the user's password in the database
    query_put("""
        UPDATE user
        SET user_password_hash = %s
        WHERE user_name = %s;
    """,
        (hashed_password, user.user_name)
    )
    user = get_user_by_username(user.user_name)
    return user[0]

# function to sort user time data so it can be used by the api
def fix_output(data):
    transformed_data = []
    for item in data:
        user_id = item['user_id']
        date = item['date']
        clock_in_time = datetime.min + item['clock_in_time']
        clock_out_time = datetime.min + item['clock_out_time']

        # Convert to 12-hour format with AM/PM indicator
        clock_in_time = clock_in_time.strftime("%I:%M:%S %p")
        clock_out_time = clock_out_time.strftime("%I:%M:%S %p")

        existing_item = next((x for x in transformed_data if x['user_id'] == user_id and x['date'] == str(date)), None)
        if existing_item:
            existing_item['clock_in_times'].append(clock_in_time)
            existing_item['clock_out_times'].append(clock_out_time)
        else:
            transformed_data.append({
                'user_id': user_id,
                'first_name': item['first_name'],
                'last_name': item['last_name'],
                'user_name': item['user_name'],
                'date': str(date),
                'clock_in_times': [clock_in_time],
                'clock_out_times': [clock_out_time],
                'total_hours': 0.0,
                'week_total_hours': 0.0
            })

    for item in transformed_data:
        # Calculate total hours
        total_seconds = sum([(datetime.strptime(out, '%I:%M:%S %p') - datetime.strptime(inp, '%I:%M:%S %p')).total_seconds() for inp, out in zip(item['clock_in_times'], item['clock_out_times'])])
        item['total_hours'] = round(total_seconds / 3600, 1)

        # Calculate week total hours
        week_items = [x for x in transformed_data if x['user_id'] == item['user_id']]
        item['week_total_hours'] = sum([x['total_hours'] for x in week_items])

    return transformed_data

def admin_get_all_users_timesheet(admin: str) -> List[dict]:

    admin = admin_get_role(admin)

    print(admin)
    # if admin is not an admin, return an error message
    if admin != 'admin':
        return "Error: You do not have permission to view this page."
    
    data = query_get_timecard("""
        SELECT 
            clock_times.user_id,
            user.user_first_name,
            user.user_last_name,
            user.user_name,
            user.user_timezone,
            DATE(clock_times.clock_in_time) AS date,
            TIME(clock_times.clock_in_time) AS clock_in_time,
            TIME(clock_times.clock_out_time) AS clock_out_time
        FROM clock_times
        JOIN user ON clock_times.user_id = user.id
    """)
    #print(data)

    results = []
    for item in data:
        user_id = item['user_id']
        first_name = item['user_first_name']
        last_name = item['user_last_name']
        user_name = item['user_name']
        user_timezone = item['user_timezone']
        date = item['date']
        clock_in_time = item['clock_in_time']
        clock_out_time = item['clock_out_time']

        # create a new result for this user and date
        result = {
            'user_id': user_id,
            'first_name': first_name,
            'last_name': last_name,
            'user_name': user_name,
            'user_timezone': user_timezone,
            'date': date,
            'clock_in_time': clock_in_time,
            'clock_out_time': clock_out_time,
        }
        results.append(result)

    # sort the results by user_id and date
    sorted_results = sorted(results, key=lambda x: (x['user_id'], x['date']))

    return fix_output(sorted_results)



# admin function to get all users' time records using AdminTimesheetResponseModelAllUsers and download as xlsx file
def admin_get_all_users_timesheet_download() -> List[dict]:

    #admin = admin_get_role(admin)

    #print(admin)
    # if admin is not an admin, return an error message
    #if admin != 'admin':
    #    return "Error: You do not have permission to view this page."
    


    data = query_get_timecard("""
        SELECT 
            clock_times.user_id,
            user.user_first_name,
            user.user_last_name,
            user.user_name,
            DATE(clock_times.clock_in_time) AS date,
            TIME(clock_times.clock_in_time) AS clock_in_time,
            TIME(clock_times.clock_out_time) AS clock_out_time
        FROM clock_times
        JOIN user ON clock_times.user_id = user.id
    """)
    #print(data)

    results = []
    for item in data:
        user_id = item['user_id']
        first_name = item['user_first_name']
        last_name = item['user_last_name']
        user_name = item['user_name']
        date = item['date']
        clock_in_time = item['clock_in_time']
        clock_out_time = item['clock_out_time']

        # create a new result for this user and date
        result = {
            'user_id': user_id,
            'first_name': first_name,
            'last_name': last_name,
            'user_name': user_name,
            'date': date,
            'clock_in_time': clock_in_time,
            'clock_out_time': clock_out_time,
        }
        results.append(result)

    # sort the results by user_id and date
    sorted_results = sorted(results, key=lambda x: (x['user_id'], x['date']))

    return fix_output(sorted_results)

# function that creates an excel file based off data from admin_get_all_users_timesheet_download
def admin_create_excel(data):
    try:
        day_names = {
            "Monday": "Mon",
            "Tuesday": "Tue",
            "Wednesday": "Wed",
            "Thursday": "Thu",
            "Friday": "Fri"
        }
        
        table_data = {
            "sales team member": [],
            "total for week": []
        }
        
        for header in day_names.values():
            table_data[f"{header} in"] = []
            table_data[f"{header} out"] = []
            table_data[f"{header} total"] = []

        for user in data:
            sales_team_member = user['first_name'] + ' ' + user['last_name']
            if sales_team_member not in table_data['sales team member']:
                table_data['sales team member'].append(sales_team_member)
                table_data['total for week'].append(0)  # initialize to 0

                for header in day_names.values():
                    table_data[f"{header} in"].append(None)
                    table_data[f"{header} out"].append(None)
                    table_data[f"{header} total"].append(None)

            try:
                current_day_of_week = datetime.strptime(user['date'], '%Y-%m-%d').strftime('%A')
                index = table_data['sales team member'].index(sales_team_member)

                if current_day_of_week in day_names.keys():
                    header = day_names[current_day_of_week]
                    in_time = datetime.strptime(user['clock_in_times'][0], '%H:%M:%S').strftime('%I:%M:%S %p')
                    out_time = datetime.strptime(user['clock_out_times'][0], '%H:%M:%S').strftime('%I:%M:%S %p')
                    total_hours = user['total_hours'] - 0.5  # subtract 30 minutes
                    table_data[f"{header} in"][index] = in_time
                    table_data[f"{header} out"][index] = out_time
                    table_data[f"{header} total"][index] = total_hours

                    # Update the total for week
                    table_data['total for week'][index] += total_hours

            except Exception as e:
                print(f"Error parsing date {user['date']}: {e}")

        # Create the DataFrame
        df = pd.DataFrame(table_data)

        # Reorder the columns to move 'total for week' to the end
        columns = ['sales team member']
        for header in day_names.values():
            columns += [f"{header} in", f"{header} out", f"{header} total"]
        columns += ['total for week']
        df = df[columns]

        # Save the DataFrame as an Excel workbook

        # Get the current date
        current_date = datetime.now()

        # Create a file name using the current date
        current_date = datetime.now()
        file_name = f"{current_date:%Y-%m-%d}.xlsx"

        # Construct the full file path using the download directory and file name
        file_path = os.path.join(DOWNLOAD_DIR, file_name)

        # Save the DataFrame to the Excel file
        writer = pd.ExcelWriter(file_path)
        df.to_excel(writer, index=False)
        writer.close()

        # Set the font styles
        header_font = Font(name='Cambria', bold=True)
        body_font = Font(name='Cambria')

        # Open the workbook and set the column width for the 'sales team member' column
        wb = load_workbook(file_path)
        ws = wb.active

        # Set the font for the first row to Cambria
        for cell in ws[1]:
            cell.font = header_font

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15

        

        # Save the updated workbook
        wb.save(file_name)
        
        return file_name
    except Exception as e:
        print(f"Error creating excel file: {e}")